from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_DIR = Path("Week7_Practice_Files")
OUTPUT_DIR.mkdir(exist_ok=True)

# ------------------------------------------------------------
# 01_Messy_Trades.csv
# Purpose: messy source file for import, cleaning, type conversion,
# and splitting Product Code into Region/Product/Year/Sequence.
# ------------------------------------------------------------
messy_trades = [
    ["Trade ID", "Trade Date", "Product Code", "Counterparty", "Quantity", "Price", "Trader", "Notes"],
    [" T-1001 ", "2026-03-01", "APAC-BUTTER-2026-001", "  Sunrise Foods Pte Ltd ", "1,200", "$4.25", " Xinran ", "urgent shipment"],
    ["T-1002", "03/02/2026", "EMEA-CHEESE-2026-002", "alpine dairy GmbH", "850", "3.80", "Maya", " confirm docs "],
    ["T-1003", "2-Mar-26", "APAC-MILKPOWDER-2026-003", "  BLUE HARBOUR TRADING", "2,500", "$2.15", "Leo", "missing PO"],
    [" T-1004", "2026/03/04", "AMER-COCOA-2026-004", "North Star Foods", "1,000", "5.60", " Xinran", ""],
    ["T-1005 ", "05-03-2026", "APAC-BUTTER-2026-005", "green valley ltd ", " 700 ", "$4.40", "Maya ", "price revised"],
    ["T-1006", "March 6 2026", "EMEA-MILKPOWDER-2026-006", "  Euro Dairy Partners", "1,450", "2.30", "Leo", ""],
    ["T-1007", "2026.03.07", "APAC-CHEESE-2026-007", "Lotus Trading Co", "920", "$3.95", "Xinran", "check credit limit"],
    [" T-1008 ", "8/3/2026", "AMER-BUTTER-2026-008", "  Golden Farm Inc ", "1,100", "4.10", "Maya", ""],
    ["T-1009", "2026-03-09", "APAC-COCOA-2026-009", "Merlion Commodities", "600", "$5.75", "Leo", "contract pending"],
    ["T-1010", "10-Mar-26", "EMEA-BUTTER-2026-010", "  Nordic Foods AS", "1,300", "4.05", "Xinran", ""],
]

with open(OUTPUT_DIR / "01_Messy_Trades.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerows(messy_trades)

# ------------------------------------------------------------
# 02_New_Trades_To_Append.csv
# Purpose: new file to append to the cleaned/current workbook.
# Same final intended structure as source file.
# ------------------------------------------------------------
new_trades = [
    ["Trade ID", "Trade Date", "Product Code", "Counterparty", "Quantity", "Price", "Trader", "Notes"],
    ["T-1011", "2026-03-11", "APAC-BUTTER-2026-011", "Harbour Foods Pte Ltd", "900", "4.35", "Xinran", "new APAC butter order"],
    ["T-1012", "2026-03-12", "EMEA-CHEESE-2026-012", "Lakeside Dairy BV", "750", "3.90", "Maya", "append test"],
    ["T-1013", "2026-03-13", "APAC-MILKPOWDER-2026-013", "Pacific Ingredients", "1800", "2.25", "Leo", "append test"],
    ["T-1014", "2026-03-14", "AMER-COCOA-2026-014", "Cocoa Bridge LLC", "500", "5.95", "Xinran", "append test"],
]

with open(OUTPUT_DIR / "02_New_Trades_To_Append.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerows(new_trades)

# ------------------------------------------------------------
# 03_Product_Mapping.csv
# Purpose: optional lookup/mapping file for extension exercise.
# ------------------------------------------------------------
product_mapping = [
    ["Product", "Product Group", "Business Unit", "Default Currency"],
    ["BUTTER", "Dairy Fats", "Dairy", "USD"],
    ["CHEESE", "Dairy Ingredients", "Dairy", "USD"],
    ["MILKPOWDER", "Dairy Powders", "Dairy", "USD"],
    ["COCOA", "Cocoa Products", "Cocoa", "USD"],
]

with open(OUTPUT_DIR / "03_Product_Mapping.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerows(product_mapping)

# ------------------------------------------------------------
# Week7_Workbook.xlsx
# Purpose: starter workbook with instructions and a sample current table
# participants can use for the append and calculated field exercises.
# ------------------------------------------------------------
wb = Workbook()
ws_intro = wb.active
ws_intro.title = "Instructions"

ws_intro["A1"] = "Week 7 Practice Workbook"
ws_intro["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws_intro["A1"].fill = PatternFill("solid", fgColor="40AE59")
ws_intro["A1"].alignment = Alignment(horizontal="center")
ws_intro.merge_cells("A1:F1")

instructions = [
    ["Task", "What to practice", "File to use"],
    ["1", "Import and clean data: change data types, trim text, split Product Code", "01_Messy_Trades.csv"],
    ["2", "Append one new file to the current Excel table", "02_New_Trades_To_Append.csv"],
    ["3", "Create calculated fields: Trade Value and Risk Flag", "Trades_Table in this workbook"],
]

for row_idx, row in enumerate(instructions, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_intro.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="40AE59")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws_intro.column_dimensions["A"].width = 12
ws_intro.column_dimensions["B"].width = 60
ws_intro.column_dimensions["C"].width = 34

ws_current = wb.create_sheet("Current_Trades")
current_headers = [
    "Trade ID", "Trade Date", "Product Code", "Counterparty", "Quantity", "Price", "Trader", "Notes"
]
current_rows = [
    ["T-0996", "2026-02-24", "APAC-BUTTER-2026-096", "Ocean Foods Pte Ltd", 1000, 4.10, "Xinran", "current workbook row"],
    ["T-0997", "2026-02-25", "EMEA-CHEESE-2026-097", "Swiss Valley AG", 650, 3.85, "Maya", "current workbook row"],
    ["T-0998", "2026-02-26", "APAC-MILKPOWDER-2026-098", "East Asia Dairy", 2200, 2.20, "Leo", "current workbook row"],
    ["T-0999", "2026-02-27", "AMER-COCOA-2026-099", "Cocoa Coast LLC", 480, 5.70, "Xinran", "current workbook row"],
    ["T-1000", "2026-02-28", "APAC-BUTTER-2026-100", "Singapore Dairy Hub", 1250, 4.18, "Maya", "current workbook row"],
]
ws_current.append(current_headers)
for row in current_rows:
    ws_current.append(row)

# Create an Excel table named Trades_Table
end_row = len(current_rows) + 1
end_col_letter = "H"
tab = Table(displayName="Trades_Table", ref=f"A1:{end_col_letter}{end_row}")
style = TableStyleInfo(
    name="TableStyleMedium4",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
tab.tableStyleInfo = style
ws_current.add_table(tab)

for col in range(1, len(current_headers) + 1):
    ws_current.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
    ws_current.cell(row=1, column=col).fill = PatternFill("solid", fgColor="40AE59")
    ws_current.column_dimensions[chr(64 + col)].width = 20

ws_current.freeze_panes = "A2"

# Blank sheet for participants to load Power Query results
ws_output = wb.create_sheet("Power_Query_Output")
ws_output["A1"] = "Load your Power Query output here."
ws_output["A1"].font = Font(italic=True, color="666666")
ws_output.column_dimensions["A"].width = 40

# Save workbook
wb.save(OUTPUT_DIR / "Week7_Workbook.xlsx")

print(f"Practice files created in: {OUTPUT_DIR.resolve()}")
print("Files:")
for file in OUTPUT_DIR.iterdir():
    print(f"- {file.name}")
